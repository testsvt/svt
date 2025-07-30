#!/usr/bin/env python3
import pandas as pd
import os
import re
import shutil
from openpyxl import load_workbook

def contains_chinese(text):
    """Проверить, содержит ли текст китайские символы"""
    if pd.isna(text) or not isinstance(text, str):
        return False
    chinese_pattern = re.compile(r'[\u4e00-\u9fff]')
    return bool(chinese_pattern.search(text))

def get_sheet_names(file_path):
    """Получить список листов в файле"""
    try:
        workbook = load_workbook(file_path, read_only=True)
        return workbook.sheetnames
    except Exception as e:
        print(f"  ❌ Ошибка при чтении листов: {e}")
        return []

def replace_chinese_in_file(source_file, target_file, column_name):
    """Заменить китайские символы в файле"""
    try:
        # Создаем резервную копию
        backup_path = target_file.replace('.xlsx', '_backup.xlsx')
        shutil.copy2(target_file, backup_path)
        print(f"  💾 Создана резервная копия: {os.path.basename(backup_path)}")
        
        # Получаем список листов
        source_sheets = get_sheet_names(source_file)
        target_sheets = get_sheet_names(target_file)
        
        print(f"  📋 Листы в исходном файле: {source_sheets}")
        print(f"  📋 Листы в целевом файле: {target_sheets}")
        
        # Находим общие листы
        common_sheets = [sheet for sheet in source_sheets if sheet in target_sheets]
        
        if not common_sheets:
            print(f"  ❌ Общие листы не найдены")
            return 0
        
        total_replaced = 0
        
        for sheet_name in common_sheets:
            try:
                print(f"  📄 Обрабатываю лист: {sheet_name}")
                
                # Читаем исходный файл (русский)
                source_df = pd.read_excel(source_file, sheet_name=sheet_name)
                
                # Читаем целевой файл (китайский)
                target_df = pd.read_excel(target_file, sheet_name=sheet_name)
                
                if column_name not in source_df.columns or column_name not in target_df.columns:
                    print(f"    ⚠️  Столбец {column_name} не найден в листе {sheet_name}")
                    continue
                
                # Создаем словарь соответствий
                mapping = {}
                min_rows = min(len(source_df), len(target_df))
                
                for i in range(min_rows):
                    chinese_name = target_df.iloc[i][column_name]
                    russian_name = source_df.iloc[i][column_name]
                    
                    if pd.notna(chinese_name) and pd.notna(russian_name):
                        if contains_chinese(str(chinese_name)):
                            mapping[str(chinese_name)] = str(russian_name)
                
                if not mapping:
                    print(f"    ⚠️  Соответствия не найдены в листе {sheet_name}")
                    continue
                
                print(f"    📝 Найдено {len(mapping)} соответствий")
                
                # Заменяем китайские символы
                replaced_count = 0
                for i, value in enumerate(target_df[column_name]):
                    if pd.notna(value) and str(value) in mapping:
                        target_df.at[i, column_name] = mapping[str(value)]
                        replaced_count += 1
                
                if replaced_count > 0:
                    # Сохраняем изменения
                    with pd.ExcelWriter(target_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        target_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    print(f"    ✅ Заменено {replaced_count} значений в листе {sheet_name}")
                    total_replaced += replaced_count
                else:
                    print(f"    ℹ️  Замены не найдены в листе {sheet_name}")
                    
            except Exception as e:
                print(f"    ❌ Ошибка при обработке листа {sheet_name}: {e}")
                continue
        
        return total_replaced
        
    except Exception as e:
        print(f"  ❌ Ошибка: {e}")
        return 0

def main():
    print("🔄 Начинаю замену китайских названий на русские...")
    print("="*60)
    
    source_dir = "Character.edf_ru"
    target_dir = "Character.edf_main"
    
    # Определяем файлы для обработки
    files_to_process = [
        {
            'source_file': os.path.join(source_dir, "Grade.xlsx"),
            'target_file': os.path.join(target_dir, "Grade.xlsx"),
            'column': 'string[32]'
        },
        {
            'source_file': os.path.join(source_dir, "Grade.xlsx"),
            'target_file': os.path.join(target_dir, "Grade.xlsx"),
            'column': 'string[32].1'
        },
        {
            'source_file': os.path.join(source_dir, "Grade.xlsx"),
            'target_file': os.path.join(target_dir, "Grade.xlsx"),
            'column': 'string[32].2'
        },
        {
            'source_file': os.path.join(source_dir, "MonsterCharacter.xlsx"),
            'target_file': os.path.join(target_dir, "MonsterCharacter.xlsx"),
            'column': 'string[64].1'
        },
        {
            'source_file': os.path.join(source_dir, "Class.xlsx"),
            'target_file': os.path.join(target_dir, "Class.xlsx"),
            'column': 'string[64].10'
        },
        {
            'source_file': os.path.join(source_dir, "Action.xlsx"),
            'target_file': os.path.join(target_dir, "Action.xlsx"),
            'column': 'string[32]'
        }
    ]
    
    total_replacements = 0
    
    for file_config in files_to_process:
        source_file = file_config['source_file']
        target_file = file_config['target_file']
        column = file_config['column']
        
        if not os.path.exists(source_file):
            print(f"❌ Исходный файл не найден: {os.path.basename(source_file)}")
            continue
            
        if not os.path.exists(target_file):
            print(f"❌ Целевой файл не найден: {os.path.basename(target_file)}")
            continue
        
        print(f"\n📁 Обрабатываю файл: {os.path.basename(target_file)}")
        print(f"   Столбец: {column}")
        
        # Заменяем китайские символы
        replaced = replace_chinese_in_file(source_file, target_file, column)
        total_replacements += replaced
    
    print(f"\n{'='*60}")
    print(f"✅ Завершено! Всего заменено {total_replacements} значений")
    print("🎉 Китайские названия успешно заменены на русские!")

if __name__ == "__main__":
    main()